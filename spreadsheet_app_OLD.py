# Give the package an alias for shorter code
import openpyxl as xl
# Import two classes (from chart module (from openpyxl package))
from openpyxl.chart import BarChart, Reference

# Returns workbook object
wb = xl.load_workbook('transactions.xlsx')
# Returns sheet object
sheet = wb['Sheet1']
# The following two lines do the same thing (just for demo, not part of program)
cell = sheet['a1']
# ..this one just uses the .cell method of sheet object
cell = sheet.cell(1,1)

# .value attribute of cell object - self-explanatory
##print(cell.value)
# Find out how many rows are in the spreadsheet
# using .max_row attribute of sheet object
##print(sheet.max_row)

# Iterate thru all rows, must put +1 at end cuz
# (as we learned) range func returns all values
# until the one BEFORE the last
# Also, start at 2 cuz 1st row is headers
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    # Create cell object for 'corrected' values
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

# Create values object to contain all values in 4th column
values = Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4)

chart = BarChart()
chart.add_data(values)
# Top left corner of chart will be at cell E2
sheet.add_chart(chart, 'e2')

# Save as new file (don't overwrite old one!)
wb.save('transactions2.xlsx')
